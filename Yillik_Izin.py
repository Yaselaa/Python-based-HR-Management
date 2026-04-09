import os
import tkinter as tk # Hatalı olan 'import tk as tk' yerine doğrusu yazıldı
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import winsound
import shutil
from datetime import datetime
import copy

# --- AYARLAR ---
DOSYA_YILLIK = "FRM-44.xlsx"
YIL = "2026"
YEDEK_KLASORU = "Excel_Yedekleri"

class YillikIzinPanel:
    def __init__(self, root):
        self.root = root
        self.root.title("Yıllık İzin Kayıt")
        self.root.geometry("500x750")
        self.root.configure(bg="#F4F6F7")

        if not os.path.exists(YEDEK_KLASORU):
            os.makedirs(YEDEK_KLASORU)

        self.sayfalar = self.sayfalari_yukle()
        self.arayuz_yap()

    def sayfalari_yukle(self):
        if os.path.exists(DOSYA_YILLIK):
            try:
                wb = load_workbook(DOSYA_YILLIK, read_only=True)
                s = wb.sheetnames
                wb.close()
                return s
            except: return []
        return []

    def yedek_al(self):
        try:
            zaman = datetime.now().strftime("%Y_%m_%d_%H_%M")
            shutil.copy2(DOSYA_YILLIK, os.path.join(YEDEK_KLASORU, f"Yedek_FRM44_{zaman}.xlsx"))
            return True
        except: return False

    def arayuz_yap(self):
        # --- STİL AYARLARI (Combobox için) ---
        style = ttk.Style()
        style.theme_use('clam')
        
        # Combobox kutusunun dışı ve içi için zorunlu renk tanımlaması
        style.map("TCombobox",
                  fieldbackground=[('readonly', "#C3C4D6")],
                  background=[('readonly', "#C3C4D6")])
                  
        style.configure("TCombobox", 
                        fieldbackground="#C3C4D6", 
                        background="#C3C4D6",
                        foreground="black")

        # Aşağı açılan listenin renkleri
        self.root.option_add("*TCombobox*Listbox.background", "#C3C4D6")
        self.root.option_add("*TCombobox*Listbox.foreground", "black")

        header = tk.Frame(self.root, bg="#06173B", height=60)
        header.pack(fill="x")
        tk.Label(header, text="İZİN KAYIT SİSTEMİ", fg="white", bg="#06173B", font=("Helvetica", 12, "bold")).pack(pady=15)

        form = tk.Frame(self.root, bg="#C3C4D6", padx=40, pady=10)
        form.pack(expand=True, fill="both")

        tk.Label(form, text="PERSONEL SEÇİMİ", bg="#C3C4D6", font=("Helvetica", 9, "bold")).pack(anchor="w")
        self.combo_sayfa = ttk.Combobox(form, values=self.sayfalar, font=("Helvetica", 11), state="readonly")
        self.combo_sayfa.pack(fill="x", pady=5)

        self.labels = ["İzin Nedeni", "Başlangıç Tarihi", "Dönüş Tarihi", "İş Başı Tarihi", "İzin Süresi"]
        self.ents = {}
        for lbl in self.labels:
            tk.Label(form, text=lbl, bg="#C3C4D6", font=("Helvetica", 9, "bold")).pack(anchor="w", pady=(10,0))
            # Textbox (Entry) renkleri #C3C4D6 olarak ayarlandı
            e = tk.Entry(form, font=("Helvetica", 11), relief="solid", bd=1, justify="center", 
                         bg="#C3C4D6", fg="black", insertbackground="black")
            e.pack(fill="x", pady=5)
            self.ents[lbl] = e

        # Enter ile hızlı tarih formatlama
        self.ents[self.labels[1]].bind('<Return>', lambda e: self.tarih_format(self.ents[self.labels[1]], self.ents[self.labels[2]]))
        self.ents[self.labels[2]].bind('<Return>', lambda e: self.tarih_format(self.ents[self.labels[2]], self.ents[self.labels[3]]))
        self.ents[self.labels[3]].bind('<Return>', lambda e: self.tarih_format(self.ents[self.labels[3]], self.ents[self.labels[4]]))

        btn = tk.Button(form, text="Yedekle & Excele Aktar", bg="#06173B", fg="white", 
                        font=("Helvetica", 11, "bold"), height=2, relief="flat", command=self.kaydet)
        btn.pack(fill="x", pady=25)

    def tarih_format(self, su_an, sonraki):
        y = su_an.get().strip()
        if " " in y:
            p = y.split(" ")
            if len(p) >= 2:
                su_an.delete(0, tk.END)
                su_an.insert(0, f"{p[0].zfill(2)}.{p[1].zfill(2)}.{YIL}")
        sonraki.focus_set()

    def kaydet(self):
        s_adi = self.combo_sayfa.get()
        if not s_adi: 
            messagebox.showwarning("Uyarı", "Personel seçilmedi!")
            return
            
        self.yedek_al()

        try:
            wb = load_workbook(DOSYA_YILLIK)
            ws = wb[s_adi]
            
            yaz_r = None
            toplam_r = None
            baslangic_r = 10 

            for r in range(baslangic_r, 1000):
                b_val = str(ws.cell(row=r, column=2).value).upper()
                if "TOPLAM" in b_val or "KALAN" in b_val:
                    toplam_r = r
                    break

            for r in range(baslangic_r, toplam_r if toplam_r else 1000):
                hucre_h = ws.cell(row=r, column=8)
                is_colored = False
                if hucre_h.fill and hucre_h.fill.start_color.index not in ['00000000', 0, 'FFFFFFFF']:
                    is_colored = True
                
                if not is_colored and (hucre_h.value is None or str(hucre_h.value).strip() == ""):
                    yaz_r = r
                    break
            
            if yaz_r is None:
                yaz_r = toplam_r
                ws.insert_rows(yaz_r)
                for c in range(1, ws.max_column + 1):
                    eski = ws.cell(row=yaz_r-1, column=c)
                    yeni = ws.cell(row=yaz_r, column=c)
                    if eski.has_style:
                        yeni.font = copy.copy(eski.font)
                        yeni.border = copy.copy(eski.border)
                        yeni.fill = PatternFill(fill_type=None)
                        yeni.alignment = copy.copy(eski.alignment)

            ws.cell(row=yaz_r, column=7, value=self.ents[self.labels[0]].get())
            ws.cell(row=yaz_r, column=8, value=self.ents[self.labels[1]].get())
            ws.cell(row=yaz_r, column=9, value=self.ents[self.labels[2]].get())
            ws.cell(row=yaz_r, column=10, value=self.ents[self.labels[3]].get())
            
            sure = self.ents[self.labels[4]].get().replace(',','.')
            try: ws.cell(row=yaz_r, column=11, value=float(sure))
            except: ws.cell(row=yaz_r, column=11, value=sure)

            wb.save(DOSYA_YILLIK)
            winsound.Beep(800, 300)
            messagebox.showinfo("Başarılı", f"Veri {yaz_r}. satıra kaydedildi.")
            
            for e in self.ents.values(): e.delete(0, tk.END)
            self.ents[self.labels[0]].focus_set()

        except Exception as e:
            messagebox.showerror("Hata", f"Kayıt sırasında hata oluştu: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = YillikIzinPanel(root)
    root.mainloop()