import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
import winsound

# --- AYARLAR ---
DOSYA_ADI = "Izin_Takip_Sistemi_Veri_Merkezi.xlsx"
HEDEF_YIL = "2026"

# --- KURUMSAL RENKLER ---
BG_COLOR = "#C3C4D6"        # Sayfa Arkaplanı (Hafif Gri)
HEADER_COLOR = "#06173B"    # Koyu Lacivert (Başlık)
ACCENT_BLUE = "#06173B"     # Kurumsal Mavi
SUCCESS_GREEN = "#28B463"   # Başarı Yeşili
ERROR_RED = "#CB4335"       # Hata Kırmızısı
TEXT_COLOR = "#2C3E50"      # Standart Yazı

class ModernEntry(tk.Entry):
    """Sade ve şık, kurumsal odaklı giriş kutusu."""
    def __init__(self, master, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.config(
            bg="white",
            fg=TEXT_COLOR,
            insertbackground=ACCENT_BLUE,
            relief="flat",
            highlightthickness=1,
            highlightbackground="#D5DBDB",
            highlightcolor=ACCENT_BLUE,
            font=("Helvetica", 11),
            justify="center"
        )

class AkilliAramaKutusu(ModernEntry):
    def __init__(self, liste, sonraki_odak, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.liste = sorted(liste)
        self.sonraki_odak = sonraki_odak
        self.eslesenler = []
        self.indeks = 0
        self.bind('<KeyRelease>', self._klavye_takip)
        self.bind('<Tab>', self._tab_ile_gezin)
        self.bind('<Return>', self._enter_ile_gec)

    def _klavye_takip(self, event):
        if event.keysym in ("Tab", "Return", "Shift_L", "Shift_R", "Control_L", "Control_R"):
            return
        yazi = self.get().lower().strip()
        if not yazi:
            self.eslesenler = []
            durum_guncelle("Bekleniyor...")
            return
        self.eslesenler = [kisi for kisi in self.liste if kisi.lower().startswith(yazi)]
        self.indeks = 0
        if self.eslesenler:
            durum_guncelle(f"Eşleşme: {self.eslesenler[0]}...", ACCENT_BLUE)

    def _tab_ile_gezin(self, event):
        if self.eslesenler:
            self.delete(0, tk.END)
            self.insert(0, self.eslesenler[self.indeks])
            self.indeks = (self.indeks + 1) % len(self.eslesenler)
            return "break"

    def _enter_ile_gec(self, event):
        if self.get():
            self.sonraki_odak.focus_set()

def tarih_formatla(event):
    gelen = tarih_ent.get().strip()
    if " " in gelen:
        parcalar = gelen.split(" ")
        if len(parcalar) >= 2:
            gun = parcalar[0].zfill(2)
            ay = parcalar[1].zfill(2)
            yeni_tarih = f"{gun}.{ay}.{HEDEF_YIL}"
            tarih_ent.delete(0, tk.END)
            tarih_ent.insert(0, yeni_tarih)
    miktar_ent.focus_set()

def personelleri_excelden_cek():
    isimler = []
    if os.path.exists(DOSYA_ADI):
        try:
            wb = load_workbook(DOSYA_ADI, data_only=True)
            ws = wb.active
            for row in range(2, ws.max_row + 1, 2):
                hucre = ws.cell(row=row, column=2).value
                if hucre: isimler.append(str(hucre).strip())
            wb.close()
        except: pass
    return isimler

def veriyi_isle(islem_tipi):
    secilen_kisi = kisi_ent.get().strip()
    tarih = tarih_ent.get().strip()
    miktar = miktar_ent.get().strip()

    if not all([secilen_kisi, tarih, miktar]):
        durum_guncelle("Eksik veri girişi!", ERROR_RED)
        return

    try:
        wb = load_workbook(DOSYA_ADI)
        ws = wb.active
        merkez = Alignment(horizontal="center", vertical="center")
        
        target_row = None
        for row in range(2, ws.max_row + 1, 2):
            if str(ws.cell(row=row, column=2).value).strip().lower() == secilen_kisi.lower():
                target_row = row
                break
        
        if not target_row:
            durum_guncelle("Personel bulunamadı!", ERROR_RED)
            return

        start_col, end_col = (4, 22) if islem_tipi == "TS" else (23, 42)
        
        eklendi = False
        for col in range(start_col, end_col):
            if ws.cell(row=target_row, column=col).value is None:
                ws.cell(row=target_row, column=col, value=tarih).alignment = merkez
                try:
                    val = float(miktar.replace(',', '.'))
                    ws.cell(row=target_row + 1, column=col, value=val).alignment = merkez
                except:
                    ws.cell(row=target_row + 1, column=col, value=miktar).alignment = merkez
                
                try:
                    wb.save(DOSYA_ADI)
                    winsound.Beep(1000, 200)
                    durum_guncelle(f"Kaydedildi: {secilen_kisi}", SUCCESS_GREEN)
                    kisi_ent.delete(0, tk.END)
                    miktar_ent.delete(0, tk.END)
                    kisi_ent.focus_set()
                except PermissionError:
                    durum_guncelle("Dosya açık, kaydedilemedi!", ERROR_RED)
                eklendi = True
                break
        
        if not eklendi:
            durum_guncelle("Sütunlar dolu!", ERROR_RED)
            
    except Exception as e:
        durum_guncelle(f"Sistem Hatası: {str(e)}", ERROR_RED)

def durum_guncelle(metin, renk="#666"):
    durum_lbl.config(text=metin, fg=renk)

# --- ARAYÜZ ---
root = tk.Tk()
root.title("İzin Takip Sistemi v10.0")
root.geometry("480x550")
root.configure(bg=BG_COLOR)

p_listesi = personelleri_excelden_cek()

# Başlık Paneli
header_frame = tk.Frame(root, bg=HEADER_COLOR, height=60)
header_frame.pack(fill="x")
tk.Label(header_frame, text="PERSONEL İZİN OTOMASYONU", font=("Helvetica", 12, "bold"), bg=HEADER_COLOR, fg="white").pack(pady=15)

# Form Alanı
form_frame = tk.Frame(root, bg=BG_COLOR, padx=40, pady=20)
form_frame.pack(expand=True, fill="both")

tk.Label(form_frame, text="Ad Soyad", bg=BG_COLOR, fg=TEXT_COLOR, font=("Helvetica", 9, "bold")).pack(anchor="w")
tarih_ent = ModernEntry(form_frame)
kisi_ent = AkilliAramaKutusu(p_listesi, tarih_ent, form_frame)
kisi_ent.pack(fill="x", pady=(5,15))
kisi_ent.focus_set()

tk.Label(form_frame, text="İzin Tarihi (G G A A)", bg=BG_COLOR, fg=TEXT_COLOR, font=("Helvetica", 9, "bold")).pack(anchor="w")
tarih_ent.pack(fill="x", pady=(5,15))
tarih_ent.bind('<Return>', tarih_formatla)

tk.Label(form_frame, text="Miktar (Saat/Gün)", bg=BG_COLOR, fg=TEXT_COLOR, font=("Helvetica", 9, "bold")).pack(anchor="w")
miktar_ent = ModernEntry(form_frame)
miktar_ent.pack(fill="x", pady=(5,20))
miktar_ent.bind('<Return>', lambda e: veriyi_isle("TS"))

durum_lbl = tk.Label(form_frame, text="Hazır", bg=BG_COLOR, fg="#888", font=("Helvetica", 9, "italic"))
durum_lbl.pack(pady=5)

# Butonlar
btn_style = {"font": ("Helvetica", 10, "bold"), "height": 2, "relief": "flat", "cursor": "hand2"}

btn_ts = tk.Button(form_frame, text="İzin Girişi Yap (Enter)", bg=ACCENT_BLUE, fg="white", **btn_style, command=lambda: veriyi_isle("TS"))
btn_ts.pack(fill="x", pady=5)

btn_is = tk.Button(form_frame, text="İşbaşı Kağıdı İşle", bg="#636B96", fg="white", **btn_style, command=lambda: veriyi_isle("IS"))
btn_is.pack(fill="x", pady=10)

# Alt Bilgi
footer = tk.Label(root, text=f"© 2026 Kurumsal Veri Yönetimi | {len(p_listesi)} Personel", bg=BG_COLOR, fg="#ABB2B9", font=("Helvetica", 8))
footer.pack(side="bottom", pady=10)

root.mainloop()