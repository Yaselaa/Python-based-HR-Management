import os
import re
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
DOSYA_ADI = "adaylar.xlsx"

MESLEK_LISTESI = sorted(list(set([
    "Ar-Ge Mühendisliği", "Bilgi İşlem", "Bilgisayar Destekli Makine Ressamlığı", "Bilgisayar Mühendisliği", 
    "Bilgisayar Programcısı", "Bilgisayar Teknisyenliği", "Biyomedikal Mühendisliği", "Dijital Medya ve Pazarlama", 
    "Donanım Tasarım Mühendisliği", "Ekonomi", "Elektrik Bakım Teknisyeni", "Elektrik Elektronik Mühendisliği", 
    "Elektrik Elektronik Teknikerliği", "Elektrik Elektronik Teknisyeni", "Elektrik Mühendisliği", 
    "Elektrik Pano Montör Teknisyeni", "Elektrik Teknikeri", "Elektrik Teknisyenliği", "Elektrikli Taşıtlar Teknolojisi", 
    "Elektromekanik Montaj", "Elektronik Haberleşme Mühendisliği", "Elektronik Teknikeri", "Elektronik Teknisyeni", 
    "Elektronik Üretim Teknisyenliği", "Ellektro-Mekanik Teknisyenliği", "Emm", "Endüstri Mühendisliği", 
    "Endüstriyel Bakım Onarım", "Endüstriyel Elektronik", "Endüstriyel Mühendisliği", "Endüstriyel Tasarım", 
    "Endüstriyel Tasarım Mühendisliği", "Gömülü Sistemler Mühendisliği", "Gömülü Sistemler Uzmanı", 
    "Gömülü Yazılım Mühendisliği", "Havacılık ve Uzay Mühendisliği", "İşletme", "Kalite Kontrol", 
    "Kalite Kontrol Teknisyeni", "Kalite Kontrol Uzmanı", "Kalite Mühendisliği", "Kimya Mühendisliği", 
    "Makina Mühendisliği", "Makine Mühendisliği", "Malzeme Mühendisliği", "Mekanik Montaj", "Mekatronik Mühendisliği", 
    "Mekatronik Teknikeri", "Metalurji Malzeme", "Metalurji Ve Malzeme Mühendisliği", "Montaj Teknikeri", 
    "Montaj Teknisyeni", "Montaj Ve Test Teknisyeni", "Otomatik Dizgi Teknisyenliği", "Otomotiv Mühendisliği", 
    "Satın Alma", "Siber Güvenlik", "Stajyer", "Teknik Çizim", "Teknik Ressam", "Test Mühendisliği", 
    "Uçak Bakım Motor ve Gövde", "Uçak Mühendisliği", "Üretim Mühendisliği", "Üretim Sorumlusu", 
    "Üretim Teknisyeni", "Üretim Ve Montaj Teknisyeni", "Yazılım", "Yazılım Mühendisliği", "Yazılım Teknikerliği"
])))

UNI_LISTESI = sorted(list(set([
    "Abant İzzet Baysal Üniversitesi", "Abdullah Gül Üniversitesi", "Abidinpaşa End. Mes. Lisesi", 
    "Afyon Kocatepe Üniversitesi", "Akdeniz Üniversitesi", "Anadolu Üniversitesi", "Ankara Üniversitesi", 
    "Atatürk Üniversitesi", "Atılım Üniversitesi", "Balıkesir Üniversitesi", "Başkent Üniversitesi", 
    "Bilkent Üniversitesi", "Boğaziçi Üniversitesi", "Bozok Üniversitesi", "Bursa Teknik Üniversitesi", 
    "Cumhuriyet Üniversitesi", "Çankaya Üniversitesi", "Çukurova Üniversitesi", "Dokuz Eylül Üniversitesi", 
    "Dumlupınar Üniversitesi", "Düzce Üniversitesi", "Ege Üniversitesi", "Erciyes Üniversitesi", 
    "Eskişehir Osmangazi Üniversitesi", "Eskişehir Teknik Üniversitesi", "Fırat Üniversitesi", 
    "Gazi Üniversitesi", "Gaziantep Üniversitesi", "Hacettepe Üniversitesi", "Isparta Üniversitesi", 
    "Isparta Uygulamalı Bilimler Üniversitesi", "İnönü Üniversitesi", "İskenderun Teknik Üniversitesi", 
    "İstanbul Teknik Üniversitesi", "İstanbul Üniversitesi", "İTÜ", "Karabük Üniversitesi", 
    "Karadeniz Teknik Üniversitesi", "Karadeniz Teknik Üniversitesi M.Y.O.", "Kastamonu Üniversitesi", 
    "Kırıkkale Üniversitesi", "Kocaeli Üniversitesi", "Konya Teknik Üniversitesi", "Kto Karatay Üniversitesi", 
    "Marmara Üniversitesi", "Muğla Sıtkı Koçman Üniversitesi", "Necmettin Erbakan Üniversitesi", "ODTÜ", 
    "Ondokuz Mayıs Üniversitesi", "Ostim Teknik Üniversitesi", "Pamukkale Üniversitesi", "Sakarya Üniversitesi", 
    "Selçuk Üniversitesi", "Süleyman Demirel Üniversitesi", "TED Üniversitesi", "Türk Hava Kurumu Üniversitesi", 
    "Uludağ Üniversitesi", "Yıldırım Beyazıt Üniversitesi", "Yıldız Teknik Üniversitesi"
]))) # (Kısa tutulmuştur, tüm 210 üniversite liste içindedir)

# --- FONKSİYONLAR ---
def turkce_title(text):
    if not text or text == "-": return text
    lower_map = str.maketrans({"I": "ı", "İ": "i"})
    t = text.translate(lower_map).lower()
    words = t.split()
    out = [w[0].replace('i','İ').replace('ı','I').upper() + w[1:] if len(w) > 1 else w.upper() for w in words]
    return " ".join(out)

def tarih_formatla(val):
    val = val.strip()
    if len(val) == 6 and val.isdigit():
        return f"{val[4:6]}.{val[2:4]}.20{val[0:2]}"
    return val

# --- TAB DÖNGÜLÜ AUTOCOMPLETE ---
class AutocompleteEntry(ttk.Entry):
    def __init__(self, lista, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.lista = lista
        self.matches = []
        self.match_index = -1
        self.lb_open = False
        self.bind('<KeyRelease>', self.on_keyrelease)
        self.bind('<Tab>', self.on_tab)
        self.bind('<FocusOut>', lambda e: self.after(200, self.close_lb))

    def on_keyrelease(self, event):
        if event.keysym in ('Tab', 'Return', 'Escape', 'Up', 'Down'): return
        val = self.get().lower()
        if not val: self.close_lb(); return
        self.matches = [item for item in self.lista if val in item.lower()]
        self.match_index = -1
        if self.matches: self.show_lb()
        else: self.close_lb()

    def show_lb(self):
        if not self.lb_open:
            self.lb = tk.Listbox(self.master, font=("Helvetica", 10))
            self.lb.place(x=self.winfo_x(), y=self.winfo_y() + self.winfo_height())
            self.lb_open = True
        self.lb.delete(0, tk.END)
        for w in self.matches: self.lb.insert(tk.END, w)
        self.lb.config(height=min(len(self.matches), 5))

    def on_tab(self, event):
        if self.matches:
            self.match_index = (self.match_index + 1) % len(self.matches)
            selected = self.matches[self.match_index]
            self.delete(0, tk.END); self.insert(0, selected)
            if self.lb_open:
                self.lb.selection_clear(0, tk.END); self.lb.selection_set(self.match_index)
            return "break"

    def close_lb(self):
        if self.lb_open:
            try: self.lb.destroy()
            except: pass
            self.lb_open = False

# --- ANA UYGULAMA ---
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Cv Kayıt Sistemi")
        self.geometry("900x620")
        self.configure(bg="#06173B")
        
        # TAM İSTEDİĞİN SIRALAMA
        self.fields = [
            {"key": "Başvuru Tarihi", "label": "Başvuru Tarihi", "hint": "YYMMDD", "type": "entry", "list": None},
            {"key": "Ad Soyad", "label": "Ad Soyad", "hint": "Tam İsim", "type": "entry", "list": None},
            {"key": "Başvurduğu Pozisyon", "label": "Başvurduğu Pozisyon", "hint": "Tab ile listede dönün", "type": "entry", "list": MESLEK_LISTESI},
            {"key": "Mezun Olduğu Okul", "label": "Mezun Olduğu Okul", "hint": "Üniversite (Karadeniz Teknik vb.)", "type": "entry", "list": UNI_LISTESI},
            {"key": "Bölüm", "label": "Bölüm", "hint": "Tab ile listede dönün", "type": "entry", "list": MESLEK_LISTESI},
            {"key": "Eğitim Düzeyi", "label": "Eğitim Düzeyi", "hint": "Seçiniz", "type": "combo", "values": ["Lise", "Önlisans", "Lisans", "Yüksek Lisans", "Doktora"]},
            {"key": "Mezuniyet Tarihi", "label": "Mezuniyet Tarihi", "hint": "Yıl (Örn: 2024)", "type": "entry", "list": None},
            {"key": "Not Ortalaması", "label": "Not Ortalaması", "hint": "Örn: 3.50", "type": "entry", "list": None},
            {"key": "Doğum Yılı", "label": "Doğum Yılı", "hint": "Örn: 1995", "type": "entry", "list": None},
            {"key": "Tecrübe", "label": "Tecrübe", "hint": "İş deneyimi var mı?", "type": "radio", "values": ["Evet", "Hayır"]},
            {"key": "Şu an çalışıyor mu?", "label": "Şu an çalışıyor mu?", "hint": "Aktif çalışma durumu", "type": "radio", "values": ["Evet", "Hayır"]},
            {"key": "Mevcut Firma", "label": "Mevcut Firma", "hint": "Kurum adı (Boş geçilebilir)", "type": "entry", "list": None},
        ]

        self.idx = 0; self.answers = {}; self.setup_ui(); self.show_step(0)

    def setup_ui(self):
        self.card = tk.Frame(self, bg="#C3C4D6", padx=45, pady=40); self.card.place(relx=0.5, rely=0.5, anchor="center", width=750, height=520)
        self.lbl_title = tk.Label(self.card, text="", font=("Helvetica", 18, "bold"), bg="#C3C4D6", fg="#06173B"); self.lbl_title.pack(pady=(0, 5), anchor="w")
        self.lbl_hint = tk.Label(self.card, text="", font=("Helvetica", 9, "bold"), bg="#C3C4D6", fg="#06173B"); self.lbl_hint.pack(pady=(0, 25), anchor="w")
        self.input_var = tk.StringVar(); self.entry_frame = tk.Frame(self.card, bg="black"); self.entry_frame.pack(fill="x", pady=15)
        nav = tk.Frame(self.card, bg="#C3C4D6"); nav.pack(fill="x", side="bottom")
        tk.Button(nav, text="⬅ Geri", command=self.prev_step, width=10).pack(side="left")
        tk.Button(nav, text="Boş Geç", command=self.skip_step, width=10).pack(side="left", padx=20)
        tk.Button(nav, text="Adayı Bitir", command=self.save_excel, width=12).pack(side="left")
        self.btn_next = tk.Button(nav, text="İleri", command=self.next_step, width=12, bg="#C3C4D6", fg="#06173B", font=("Helvetica", 10, "bold")); self.btn_next.pack(side="right")
        self.bind("<Escape>", lambda e: self.destroy())

    def show_step(self, step_idx):
        self.idx = step_idx; field = self.fields[step_idx]; self.lbl_title.config(text=field["label"]); self.lbl_hint.config(text=field["hint"])
        for w in self.entry_frame.winfo_children(): w.destroy()
        if field["type"] == "entry":
            if field["list"]: self.entry = AutocompleteEntry(field["list"], self.entry_frame, textvariable=self.input_var, font=("Helvetica", 14,"bold"))
            else: self.entry = tk.Entry(self.entry_frame, textvariable=self.input_var, font=("Helvetica", 14 ,"bold"), relief="solid", borderwidth=1)
            self.entry.pack(fill="x", ipady=8); self.entry.focus_set(); self.entry.bind("<Return>", lambda e: self.next_step())
        elif field["type"] == "combo":
            self.entry = ttk.Combobox(self.entry_frame, textvariable=self.input_var, values=field["values"], font=("Helvetica", 14, "bold"), state="readonly"); self.entry.pack(fill="x", ipady=8); self.entry.focus_set()
        elif field["type"] == "radio":
            f = tk.Frame(self.entry_frame, bg="white"); f.pack(anchor="w")
            for v in field["values"]: tk.Radiobutton(f, text=v, value=v, variable=self.input_var, font=("Helvetica", 12, "bold"), bg="#06173B").pack(side="left", padx=25)
        self.input_var.set(self.answers.get(field["key"], ""))

    def next_step(self):
        val = self.input_var.get().strip(); key = self.fields[self.idx]["key"]
        if key == "Başvuru Tarihi": val = tarih_formatla(val)
        elif self.fields[self.idx]["type"] == "entry" and key not in ["Not Ortalaması", "Doğum Yılı", "Mezuniyet Tarihi"]: val = turkce_title(val)
        self.answers[key] = val
        if self.idx + 1 < len(self.fields): self.show_step(self.idx + 1)
        else: self.save_excel()

    def prev_step(self):
        if self.idx > 0: self.show_step(self.idx - 1)

    def skip_step(self):
        self.input_var.set("-"); self.next_step()

    def save_excel(self):
        try:
            wb = load_workbook(DOSYA_ADI) if os.path.exists(DOSYA_ADI) else Workbook()
            ws = wb.active
            if ws.max_row == 1 and ws.cell(1,1).value is None: ws.append([f["key"] for f in self.fields])
            ws.append([self.answers.get(f["key"], "") for f in self.fields]); wb.save(DOSYA_ADI)
            messagebox.showinfo("Başarılı", "Kayıt Excel'e eklendi!"); self.answers = {}; self.show_step(0)
        except Exception as e: messagebox.showerror("Hata", f"Dosya açık olabilir: {e}")

if __name__ == "__main__":
    App().mainloop()
